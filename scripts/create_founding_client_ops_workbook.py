#!/usr/bin/env python3
"""Create the founding-client operations workbook."""
from __future__ import annotations

from pathlib import Path
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

NAVY = "172033"
YELLOW = "F4C542"
LIGHT = "F3F5F8"
RED = "FDE8E8"
GREEN = "E7F6EC"

SHEETS = {
    "Clients": [
        "client_id", "business_name", "service_start_date", "renewal_date",
        "client_status", "primary_contact_name", "lead_endpoint_status",
        "billing_status", "fully_paid_months", "export_eligible_date",
        "cancellation_requested_at", "final_service_date", "ghl_location_id", "notes",
    ],
    "Lead Outcomes": [
        "lead_id", "submission_id", "client_id", "received_at", "source",
        "service", "city", "contact_status", "first_contacted_at",
        "appointment_status", "outcome", "opportunity_value", "won_revenue",
        "consent_email_at", "consent_sms_at", "last_verified_at", "notes",
    ],
    "Monthly Reports": [
        "client_id", "reporting_month", "leads_received", "leads_contacted",
        "appointments_booked", "jobs_won", "jobs_lost", "outcome_not_confirmed",
        "client_confirmed_won_revenue", "owner_email_failures",
        "acknowledgment_failures", "priority_improvement", "priority_owner",
        "target_date", "client_confirmed_by", "client_confirmed_at",
        "report_sent_at", "report_path", "notes",
    ],
    "Offboarding": [
        "client_id", "request_received_at", "requester", "authorized",
        "renewal_date", "final_service_date", "billing_action",
        "billing_verified_at", "endpoint_action", "endpoint_verified_at",
        "export_status", "retention_decision", "final_confirmation_sent_at",
        "evidence_path", "notes",
    ],
    "GHL Field Map": [
        "booked_out_field", "interim_source", "ghl_object", "ghl_field",
        "required", "migration_note",
    ],
}

VALIDATIONS = {
    "Clients": {
        "client_status": ["prospect", "onboarding", "active", "cancel_pending", "ended"],
        "lead_endpoint_status": ["disabled", "enabled", "rotated", "revoked"],
        "billing_status": ["not_started", "active", "past_due", "cancel_scheduled", "cancelled"],
    },
    "Lead Outcomes": {
        "contact_status": ["new", "attempted", "contacted", "unreachable"],
        "appointment_status": ["not_set", "scheduled", "completed", "cancelled", "no_show"],
        "outcome": ["open", "won", "lost", "not_confirmed"],
    },
    "Offboarding": {
        "authorized": ["yes", "no", "pending"],
        "export_status": ["not_requested", "not_yet_eligible", "eligible", "in_progress", "delivered", "declined"],
        "retention_decision": ["retain_per_policy", "legal_hold", "deletion_requested", "pending"],
    },
}

GHL_ROWS = [
    ["client_id", "client onboarding UUID", "Location", "Booked Out Client ID", "yes", "Stable tenant key"],
    ["lead_id", "client lead UUID", "Contact", "Booked Out Lead ID", "yes", "Unique external key"],
    ["submission_id", "public intake UUID", "Contact", "Submission ID", "yes", "Deduplication correlation"],
    ["source", "website/form identifier", "Contact", "Contact Source", "yes", "Preserve original source"],
    ["contact_status", "operations workbook", "Opportunity", "Pipeline Stage", "yes", "Map deliberately"],
    ["appointment_status", "client confirmation", "Opportunity", "Appointment Status", "no", "Do not infer"],
    ["outcome", "client confirmation", "Opportunity", "Status/Lost Reason", "yes", "Unknown stays unknown"],
    ["opportunity_value", "client estimate", "Opportunity", "Monetary Value", "no", "Estimate only"],
    ["won_revenue", "client-confirmed value", "Opportunity", "Won Value", "no", "Client-confirmed only"],
    ["consent_email_at", "intake evidence", "Contact", "Email Consent Timestamp", "no", "Evidence, not assumption"],
    ["consent_sms_at", "affirmative SMS evidence", "Contact", "SMS Consent Timestamp", "no", "Never derive from email consent"],
]


def style_sheet(ws, headers):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(1, len(headers)).coordinate}"
    ws.row_dimensions[1].height = 28
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for idx, header in enumerate(headers, 1):
        width = max(14, min(32, len(header) + 4))
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = width
    for row in range(2, 502):
        if row % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=LIGHT)


def add_validations(ws, headers, rules):
    for header, values in rules.items():
        col = headers.index(header) + 1
        formula = '"' + ",".join(values) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Choose a value from the approved list."
        dv.errorTitle = "Invalid status"
        ws.add_data_validation(dv)
        dv.add(f"{ws.cell(2, col).column_letter}2:{ws.cell(501, col).column_letter}501")


def create_workbook(path: Path) -> None:
    wb = Workbook()
    instructions = wb.active
    if instructions is None:
        raise RuntimeError("Workbook did not create an active worksheet")
    instructions.title = "Instructions"
    instructions.sheet_view.showGridLines = False
    rows = [
        ["Booked Out — Founding Client Operations"],
        ["Purpose", "Operate the first one or two clients without building a temporary CRM."],
        ["PII rule", "Keep email, phone, and inquiry details in Booked Out. Use immutable IDs and operational outcomes here."],
        ["Truth rule", "Unknown and client-unconfirmed values remain unknown; never record them as zero, lost, or completed."],
        ["Daily", "Reconcile new lead IDs, contact status, appointments, and outcomes."],
        ["Weekly", "Reconcile counts with Website Leads and review open/unconfirmed outcomes."],
        ["Monthly", "Freeze a reporting month and complete the monthly evidence report template."],
        ["Migration", "Move to GHL after two clients, 25–50 leads/month, >2 manual hours/week, or a missed lead."],
    ]
    for row in rows:
        instructions.append(row)
    instructions["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    instructions["A1"].fill = PatternFill("solid", fgColor=NAVY)
    instructions.merge_cells("A1:B1")
    instructions.column_dimensions["A"].width = 20
    instructions.column_dimensions["B"].width = 105
    for row in range(2, len(rows) + 1):
        instructions[f"A{row}"].font = Font(bold=True)
        instructions[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")

    for name, headers in SHEETS.items():
        ws = wb.create_sheet(name)
        ws.append(headers)
        if name == "GHL Field Map":
            for row in GHL_ROWS:
                ws.append(row)
        style_sheet(ws, headers)
        add_validations(ws, headers, VALIDATIONS.get(name, {}))

    leads = wb["Lead Outcomes"]
    outcome_col = SHEETS["Lead Outcomes"].index("outcome") + 1
    outcome_letter = get_column_letter(outcome_col)
    leads.conditional_formatting.add(
        f"{outcome_letter}2:{outcome_letter}501",
        FormulaRule(formula=[f'${outcome_letter}2="won"'], fill=PatternFill("solid", fgColor=GREEN)),
    )
    leads.conditional_formatting.add(
        f"{outcome_letter}2:{outcome_letter}501",
        FormulaRule(formula=[f'${outcome_letter}2="not_confirmed"'], fill=PatternFill("solid", fgColor=RED)),
    )

    for sheet in ["Lead Outcomes", "Monthly Reports"]:
        ws = wb[sheet]
        for header in ["opportunity_value", "won_revenue", "client_confirmed_won_revenue"]:
            if header in SHEETS[sheet]:
                col = SHEETS[sheet].index(header) + 1
                for row in range(2, 502):
                    ws.cell(row, col).number_format = '$#,##0.00'

    wb.calculation.fullCalcOnLoad = True
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def verify_workbook(path: Path) -> None:
    wb = load_workbook(path, read_only=False, data_only=False)
    expected = ["Instructions", *SHEETS.keys()]
    if wb.sheetnames != expected:
        raise SystemExit(f"Unexpected sheets: {wb.sheetnames}")
    for name, headers in SHEETS.items():
        actual = [wb[name].cell(1, i).value for i in range(1, len(headers) + 1)]
        if actual != headers:
            raise SystemExit(f"Header mismatch in {name}")
    if not wb["Lead Outcomes"].data_validations.dataValidation:
        raise SystemExit("Lead Outcomes validations missing")


def main() -> int:
    output = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("docs/operations/Booked-Out-Founding-Client-Ops.xlsx")
    create_workbook(output)
    verify_workbook(output)
    print(f"created {output} ({output.stat().st_size} bytes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
